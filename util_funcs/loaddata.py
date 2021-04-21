import pickle

import numpy as np
from scipy import interpolate
from openpyxl import load_workbook


def aggregate(a, n=3):
    cumsum = np.cumsum(a, dtype=float)
    ret = []
    for i in range(-1, len(a) - n, n):
        low_index = i if i >= 0 else 0
        ret.append(cumsum[low_index + n] - cumsum[low_index])
    return ret


def load(
    file_path,
    sheet_name="Sheet1",
    aggre_delta=1,
    normalize=True,
    zero_fill_method='prevlatter',
    verbose=True,
):
    """Load metric data from file_path. Each column is one variable.

    Params:
        file_path:
        sheet_name: name of sheet to load.
        normaliza: normalize data by subtract mean and divide by std.
        fill_zeros: fill 0 data with nearest sample.
        verbose: the debugging print level: 0 (Nothing), 1 (Method info), 2 (Phase info), 3(Algorithm info)

    Returns:
        data     : data in numpy array format, shape of [T, N], each column is a variable
        data_head: service names
    """
    # verbose >= 3, print data loading info
    if verbose and verbose >= 3:
        print("{:^10}{:<30}:".format("", "Data path"), file_path)
    # region Read excel sheet, each row of data is one variable
    wb = load_workbook(file_path, read_only=True)
    sheet = wb[sheet_name]
    data = []
    data_head = []
    for row_values in sheet.iter_rows(
        min_row=1, max_row=sheet.max_row, max_col=sheet.max_column, values_only=True
    ):
        data_head.append(row_values[0])
        data.append(row_values[1:])

    if verbose and verbose >= 3:
        print("{:^10}{:<30}: ".format("", "Sheet Names"), end="")
        for name in wb.sheetnames:
            print(name, end=", ")
        print("")
        print("{:^10}{:<30}:".format("", "Loaded Sheet"), sheet_name)
        print(
            "{:^10}{:<30}:".format("", "Sheet Size"),
            "{} x {}".format(sheet.max_row, sheet.max_column),
        )
    wb.close()
    # endregion

    # region Aggregate data
    if aggre_delta != 1:
        # Aggregate data
        data = [aggregate(row, aggre_delta) for row in data]
    # transpose data, now each column is one variable
    data = np.array(data).T
    if verbose and verbose >= 3:
        print("{:^10}{:<30}:".format("", "Aggregate delta"), aggre_delta)
        print("{:^10}{:<30}:".format("", "Data Shape"), data.shape)
    # endregion

    zero_count = np.sum(data == 0, axis=0)
    # Fill 0s in data
    if zero_fill_method == 'prevlatter':
        if verbose:
            print("{:^10}{:<30}:".format("", "Zero fill method"), "Previous then latter")
        filled_data = data.copy()
        for j in range(filled_data.shape[1]):
            for i in range(filled_data.shape[0]):
                if filled_data[i, j] == 0 and i >= 1:
                    filled_data[i, j] = filled_data[i - 1, j]
        for j in range(filled_data.shape[1] - 1, -1, -1):
            for i in range(filled_data.shape[0] - 1, -1, -1):
                if filled_data[i, j] == 0 and i <= filled_data.shape[0] - 2:
                    filled_data[i, j] = filled_data[i + 1, j]
        data = filled_data
    elif zero_fill_method in ['linear', 'nearest', 'zero', 'slinear', 'quadratic', 'cubic', 'previous', 'next']:
        # Possible interpolate methods are:
        # linear, nearest, zero, slinear, quadratic, cubic, previous, next
        if verbose:
            print(
                "{:^10}{:<30}:".format("", "Zero fill method"), zero_fill_method + " interpolate"
            )
        x = np.arange(data.shape[0])
        new_data = []
        for var in range(data.shape[1]):
            ind = data[:, var].nonzero()
            f = interpolate.interp1d(
                x[ind[0]],
                data[ind[0], var],
                kind=zero_fill_method,
                fill_value="extrapolate",
            )
            new_data.append(f(x))
        data = np.array(new_data).T

    # Normalize data by subtract mean and divide by std
    if normalize:
        data_mean = np.mean(data, axis=0, keepdims=True)
        data_std = np.std(data, axis=0, keepdims=True)
        data = (data - data_mean) / data_std

    # print data attributes
    if verbose:
        print("{:^10}{:<30}:".format("", "Data header"))
        for i, head in enumerate(data_head):
            print("{:>15}({:4d} 0s):{}".format(i + 1, zero_count[i], head))
    return data, data_head
