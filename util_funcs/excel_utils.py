import os
import sys

from openpyxl import load_workbook
from openpyxl import Workbook


def readExl(fileName):
    """Load data from Excel
    Params:
        fileName:
    Return:
        data matrix in list format
    """
    if not os.path.exists(fileName):
        print('File {} not exists!'.format(fileName))
        sys.exit(1)
    wb = load_workbook(fileName)
    ws = wb.get_sheet_by_name('Sheet1')
    nrows = len(list(ws.rows))
    for row in ws.rows:
        ncols = len(list(row))
        break
    ws_rows_len = nrows  # 行数
    ws_columns_len = ncols  # 列数
    # 分行
    a = [[0 for col in range(ws_columns_len)] for row in range(ws_rows_len)]
    for column in range(1, ws_columns_len + 1):
        for row in range(1, ws_rows_len + 1):
            a[row - 1][column - 1] = ws.cell(row=row, column=column).value
    return a


def saveToExcel(fileName, a):
    """Save to Excel

    Params:
        fileName:
        a:
    """
    # print('Saving to Excel:', fileName)
    w = Workbook()
    ws = w.create_sheet()
    ilen = len(a)
    jlen = len(a[0])
    for i in range(ilen):
        for j in range(jlen):
            ws.cell(row=i+1, column=j+1).value = a[i][j]
    if not os.path.exists(os.path.dirname(fileName)):
        os.makedirs(os.path.dirname(fileName))
    w.save(fileName)
